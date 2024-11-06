from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from webscrapper.utils import Files, print_class_name


class Excel:

    @print_class_name
    def __init__(self, book_name: str = "book") -> None:
        """
        Create a new Excel book with given name or load it if exists at package parent folder.

        Args:
            book_name (str): book to be create it or open it.
        """
        self.__name: str = book_name
        self.__path: str = Files.create_path_outside(f"{book_name}.xlsx")
        self.__workbook: Workbook

        if Files.file_exists(self.__path):
            print(f"Opening Excel book {self.__name}")
            self.__workbook = load_workbook(self.__path)
        else:
            print(f"Creating Excel book {self.__name}")
            self.__workbook = Workbook()
            if self.__workbook.active:
                self.__workbook.remove(self.__workbook.active)

    @property
    def sheets(self) -> Sequence[Worksheet]:
        return self.__workbook.worksheets

    def has_sheet(self, sheet_name: str) -> bool:
        """
        Return true if book contains sheet with given name.

        Args:
            sheet_name (str): sheet to be checked.
        Returns:
            bool
        """
        return sheet_name in self.__workbook.sheetnames

    @print_class_name
    def create_sheet(self, sheet_name: str) -> None:
        """
        Create a new sheet with given name.

        Args:
            sheet_name (str): name of sheet to be created.
        """
        print(f"Creating sheet {sheet_name}")
        self.__workbook.create_sheet(sheet_name)

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """
        Return reference to sheet with given name.

        Args:
            sheet_name (str): name of sheet to be returned.
        """
        return self.__workbook[sheet_name]

    def delete_sheet(self, sheet_name) -> None:
        """
        Delete sheet with given name.

        Args:
            sheet_name (str): name of sheet to be deleted.
        """
        self.__workbook.remove(self.__workbook[sheet_name])

    def save_book(self) -> None:
        """
        Save Excel book at path.
        """
        self.__workbook.save(self.__path)

    @print_class_name
    def close(self) -> None:
        """
        Close Excel book.
        """
        print(f"Closing Excel book {self.__name}")
        self.__workbook.close()
